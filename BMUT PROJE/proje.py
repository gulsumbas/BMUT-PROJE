import pandas as pd              # veri okuma ve tablo (DataFrame) işlemleri için kullanılır
import numpy as np               # sayısal işlemler ve matematiksel hesaplamalar için kullanılır
import matplotlib.pyplot as plt  # grafik çizmek için kullanılır

from sklearn.model_selection import train_test_split   # veriyi eğitim ve test olarak ayırmak için kullanılır
from sklearn.compose import ColumnTransformer          # farklı sütunlara farklı işlemler uygulamak için kullanılır
from sklearn.pipeline import Pipeline                  # veri ön işleme ve modeli tek bir yapı içinde birleştirir
from sklearn.preprocessing import OneHotEncoder, StandardScaler  # kategorik veriyi sayıya çevirir ve sayısal veriyi ölçeklendirir
from sklearn.impute import SimpleImputer              # eksik verileri doldurmak için kullanılır

from sklearn.linear_model import LinearRegression      # doğrusal ilişkiyi modelleyen temel regresyon algoritması
from sklearn.tree import DecisionTreeRegressor         # veriyi kurallara bölerek öğrenen ağaç tabanlı model
from sklearn.ensemble import RandomForestRegressor     # birden fazla karar ağacını birleştirerek daha güçlü model oluşturur
from sklearn.svm import SVR  # SVM algoritmasının regresyon modeli

from sklearn.metrics import mean_squared_error, mean_absolute_error, r2_score  # model performansını ölçmek için kullanılan metrikler


df = pd.read_csv("anket_verisi.csv")        # veri okuma

rename_map = {          #sütun isimlerinin kısaltılması
    "Zaman damgası": "Zaman",
    "Cinsiyetiniz nedir?": "Cinsiyet",
    "Kaçıncı sınıfta eğitim görüyorsunuz?": "Sinif",
    "Hangi fakültede öğrenim görüyorsunuz?": "Fakulte",
    "Not ortalamanız hangi sistem üzerinden hesaplanıyor?": "Not_Sistemi",
    "4'lük sistemde geçen dönemki ortalamanızı giriniz (Küsüratlı değerleri virgül yerine noktayla ayırınız. Örn: 2.15):": "Gecen_4luk",
    "4'lük sistemde genel ortalamanızı giriniz (Küsüratlı değerleri virgül yerine noktayla ayırınız. Örn: 2.15):": "Genel_4luk",
    "100'lük sistemde geçen dönemki ortalamanızı giriniz (Küsüratlı değerleri virgül yerine noktayla ayırınız. Örn: 93.5):": "Gecen_100luk",
    "100'lük sistemde genel ortalamanızı giriniz (Küsüratlı değerleri virgül yerine noktayla ayırınız. Örn: 93.5):": "Genel_100luk",
    "Önceki dönem aldığınız ders sayısını giriniz:": "Ders_Sayisi",
    "Önceki dönem başarısız olduğunuz ders sayısını giriniz:": "Basarisiz_Ders",
    "Haftada kaç saat ders çalışıyorsunuz?": "Haftalik_Calisma",
    "Sınav haftaları kaç saat ders çalışıyorsunuz?": "Sinav_Haftasi_Calisma",
    "Annenizin eğitim durumu nedir?": "Anne_Egitim",
    "Babanızın eğitim durumu nedir?": "Baba_Egitim",
    "Evinizin aylık toplam geliri ne kadardır?": "Aylik_Gelir",
    "Bir işte çalışıyor musunuz?": "Calisma_Durumu",
    "Eğitim-öğretim dönemi içerisinde nerede ikamet ediyorsunuz?": "Ikamet",
    "Günlük ortalama ekran sürenizi seçiniz:": "Ekran_Suresi",
    "Günlük ortalama uyku sürenizi seçiniz:": "Uyku_Suresi"
}

df.rename(columns=rename_map, inplace=True)

def temizle_sayi(x):
    if pd.isna(x):
        return np.nan

    s = str(x).strip()
    if s == "":
        return np.nan

    if "," in s and "." in s:
        if s.rfind(",") > s.rfind("."):
            s = s.replace(".", "")
            s = s.replace(",", ".")
        else:
            s = s.replace(",", "")
    else:
        s = s.replace(",", ".")

    try:
        return float(s)
    except:
        return np.nan

sayisal_sutunlar = [
    "Gecen_4luk",
    "Genel_4luk",
    "Gecen_100luk",
    "Genel_100luk",
    "Ders_Sayisi",
    "Basarisiz_Ders"
]

for col in sayisal_sutunlar:
    if col in df.columns:
        df[col] = df[col].apply(temizle_sayi)

# ortak olarak 100'lük sisteme çeviriyoruz başarı değişkenini
df["Ortak_Basari_100"] = np.where(
    df["Genel_100luk"].notna(),
    df["Genel_100luk"],
    df["Genel_4luk"] * 25
)

df["Ortak_Gecen_100"] = np.where(
    df["Gecen_100luk"].notna(),
    df["Gecen_100luk"],
    df["Gecen_4luk"] * 25
)

df["Ortak_Basari_4"] = np.where(
    df["Genel_4luk"].notna(),
    df["Genel_4luk"],
    df["Genel_100luk"] / 25
)

# Geçersiz değerleri temizliyoruz
df.loc[(df["Ortak_Basari_100"] < 0) | (df["Ortak_Basari_100"] > 100), "Ortak_Basari_100"] = np.nan
df.loc[(df["Ortak_Gecen_100"] < 0) | (df["Ortak_Gecen_100"] > 100), "Ortak_Gecen_100"] = np.nan
df.loc[(df["Ortak_Basari_4"] < 0) | (df["Ortak_Basari_4"] > 4), "Ortak_Basari_4"] = np.nan

# Boş olanları atıyoruz
df = df[df["Ortak_Basari_100"].notna()].copy()

# =========================================================
# 5) TEMEL BİLGİLER
# =========================================================
print("\n--- VERİ SETİ GENEL BİLGİ ---")
print(df.info())

print("\n--- TOPLAM GEÇERLİ GÖZLEM SAYISI ---")
print(len(df))

print("\n--- NOT SİSTEMİ DAĞILIMI ---")
print(df["Not_Sistemi"].value_counts(dropna=False))

# =========================================================
# 6) İSTATİSTİKSEL ANALİZ
# =========================================================
analiz_sayisal = [
    "Ortak_Basari_100",
    "Ortak_Gecen_100",
    "Ortak_Basari_4",
    "Ders_Sayisi",
    "Basarisiz_Ders"
]

analiz_sayisal = [c for c in analiz_sayisal if df[c].notna().sum() >= 2]

print("\n--- SAYISAL DEĞİŞKENLERİN TANIMSAL İSTATİSTİKLERİ ---")
print(df[analiz_sayisal].describe())

kategorik_sutunlar = [
    "Cinsiyet",
    "Sinif",
    "Fakulte",
    "Not_Sistemi",
    "Haftalik_Calisma",
    "Sinav_Haftasi_Calisma",
    "Anne_Egitim",
    "Baba_Egitim",
    "Aylik_Gelir",
    "Calisma_Durumu",
    "Ikamet",
    "Ekran_Suresi",
    "Uyku_Suresi"
]

print("\n--- KATEGORİK DEĞİŞKENLERİN FREKANSLARI ---")
for col in kategorik_sutunlar:
    print(f"\n{col}")
    print(df[col].value_counts(dropna=False))

# =========================================================
# 7) DEĞİŞKENLERİN BAŞARI ÜZERİNDEKİ ETKİSİ
# =========================================================
print("\n--- KATEGORİK DEĞİŞKENLERE GÖRE ORTALAMA BAŞARI ---")
for col in kategorik_sutunlar:
    print(f"\n{col}")
    print(df.groupby(col)["Ortak_Basari_100"].agg(["count", "mean"]).sort_values("mean", ascending=False))

print("\n--- SAYISAL DEĞİŞKENLERİN BAŞARIYLA KORELASYONU ---")
corr_target = df[analiz_sayisal].corr(numeric_only=True)["Ortak_Basari_100"].sort_values(ascending=False)
print(corr_target)

df.groupby("Fakulte")["Ortak_Basari_100"].mean().sort_values().plot(kind="bar")
plt.title("Fakülteye Göre Ortalama Başarı")
plt.ylabel("Ortalama Başarı")
plt.xticks(rotation=45)
plt.tight_layout()
plt.show()

# =========================================================
# 8) DEĞİŞKENE UYGUN OTOMATİK GRAFİK OLUŞTURMA
# =========================================================
plt.rcParams["figure.figsize"] = (10, 6)
plt.rcParams["font.size"] = 10

for column in df.columns:
    plt.figure()

    if pd.api.types.is_numeric_dtype(df[column]):
        temiz_veri = df[column].dropna()

        if len(temiz_veri) > 0:
            plt.hist(temiz_veri, bins=12, edgecolor="black")
            ortalama = temiz_veri.mean()
            medyan = temiz_veri.median()

            plt.axvline(ortalama, linestyle="--", linewidth=2, label=f"Ortalama: {ortalama:.2f}")
            plt.axvline(medyan, linestyle="-.", linewidth=2, label=f"Medyan: {medyan:.2f}")

            plt.title(f"{column} - Sayısal Dağılım", fontsize=13, fontweight="bold")
            plt.xlabel(column)
            plt.ylabel("Frekans")
            plt.grid(axis="y", linestyle="--", alpha=0.5)
            plt.legend()
        else:
            plt.text(0.5, 0.5, "Gösterilecek veri yok", ha="center", va="center")
            plt.title(f"{column} - Sayısal Dağılım")

    else:
        sayim = df[column].fillna("Boş").astype(str).value_counts()
        kategori_sayisi = len(sayim)

        if kategori_sayisi > 0:
            if kategori_sayisi <= 3:
                plt.pie(
                    sayim.values,
                    labels=sayim.index,
                    autopct="%1.1f%%",
                    startangle=90
                )
                plt.title(f"{column} - Oransal Dağılım", fontsize=13, fontweight="bold")
                plt.axis("equal")
            else:
                if kategori_sayisi > 12:
                    sayim = sayim.head(12)

                ax = sayim.plot(kind="bar", edgecolor="black")

                plt.title(f"{column} - Kategorik Dağılım", fontsize=13, fontweight="bold")
                plt.xlabel(column)
                plt.ylabel("Frekans")
                plt.xticks(rotation=35, ha="right")
                plt.grid(axis="y", linestyle="--", alpha=0.5)

                for p in ax.patches:
                    yukseklik = p.get_height()
                    ax.annotate(
                        str(int(yukseklik)),
                        (p.get_x() + p.get_width() / 2, yukseklik),
                        ha="center",
                        va="bottom",
                        fontsize=9
                    )
        else:
            plt.text(0.5, 0.5, "Gösterilecek veri yok", ha="center", va="center")
            plt.title(f"{column} - Grafik")

    plt.tight_layout()
    plt.show()

# =========================================================
# 9) KORELASYON ANALİZİ
# =========================================================
corr_matrix = df[analiz_sayisal].corr(numeric_only=True)

print("\n--- KORELASYON MATRİSİ ---")
print(corr_matrix)

plt.figure(figsize=(8, 6))
plt.imshow(corr_matrix, cmap="coolwarm", interpolation="nearest")
plt.colorbar()
plt.xticks(range(len(corr_matrix.columns)), corr_matrix.columns, rotation=45, ha="right")
plt.yticks(range(len(corr_matrix.columns)), corr_matrix.columns)
plt.title("Korelasyon Matrisi")
plt.tight_layout()
plt.show()

plt.figure(figsize=(8, 5))
df["Ortak_Basari_100"].hist(bins=10)
plt.title("Genel Başarı Dağılımı")
plt.xlabel("Başarı Notu (100'lük)")
plt.ylabel("Frekans")
plt.tight_layout()
plt.show()

# =========================================================
# 10) MODEL İÇİN VERİ HAZIRLAMA
# =========================================================
sayisal_ozellikler = [
    "Ortak_Gecen_100",
    "Ders_Sayisi",
    "Basarisiz_Ders"
]

kategorik_ozellikler = [
    "Cinsiyet",
    "Sinif",
    "Fakulte",
    "Haftalik_Calisma",
    "Sinav_Haftasi_Calisma",
    "Anne_Egitim",
    "Baba_Egitim",
    "Aylik_Gelir",
    "Calisma_Durumu",
    "Ikamet",
    "Ekran_Suresi",
    "Uyku_Suresi"
]

X = df[sayisal_ozellikler + kategorik_ozellikler].copy()
y = df["Ortak_Basari_100"].copy()

# =========================================================
# 11) EĞİTİM / TEST AYRIMI
# =========================================================
X_train, X_test, y_train, y_test = train_test_split(
    X, y, test_size=0.20, random_state=42
)

print("\n--- EĞİTİM / TEST BOYUTLARI ---")
print("X_train:", X_train.shape)
print("X_test :", X_test.shape)
print("y_train:", y_train.shape)
print("y_test :", y_test.shape)

# =========================================================
# 12) PREPROCESSOR
# =========================================================
num_transformer = Pipeline(steps=[
    ("imputer", SimpleImputer(strategy="median")),
    ("scaler", StandardScaler())
])

cat_transformer = Pipeline(steps=[
    ("imputer", SimpleImputer(strategy="most_frequent")),
    ("onehot", OneHotEncoder(handle_unknown="ignore"))
])

preprocessor = ColumnTransformer(
    transformers=[
        ("num", num_transformer, sayisal_ozellikler),
        ("cat", cat_transformer, kategorik_ozellikler)
    ]
)

# =========================================================
# 12.1) MODELİN GÖRDÜĞÜ X VE Y TABLOSU
# =========================================================

# Preprocessor'ı uygula
X_hazir = preprocessor.fit_transform(X)

# Eğer sparse matrix ise dense hale çevir
if hasattr(X_hazir, "toarray"):
    X_hazir = X_hazir.toarray()

# OneHot sonrası kategorik sütun isimleri
encoded_cat_features = preprocessor.named_transformers_["cat"]["onehot"].get_feature_names_out(kategorik_ozellikler)

# Tüm sütun isimleri
tum_feature_isimleri = sayisal_ozellikler + list(encoded_cat_features)

# DataFrame'e çevir
X_hazir_df = pd.DataFrame(X_hazir, columns=tum_feature_isimleri)

# y'yi DataFrame yap
y_df = pd.DataFrame(y, columns=["Ortak_Basari_100"])

# İlk 10 satırlık örnek tablo
X_ornek_df = X_hazir_df.head(10).copy()
y_ornek_df = y_df.head(10).copy()

# Ekrana yazdır
print("\n--- MODELİN GÖRDÜĞÜ X MATRİSİ (İLK 10 SATIR) ---")
print(X_ornek_df.to_string(index=False))

print("\nToplam gözlem sayısı:", X_hazir_df.shape[0])
print("Toplam özellik (feature) sayısı:", X_hazir_df.shape[1])

print("\n--- MODELİN GÖRDÜĞÜ Y (HEDEF DEĞİŞKEN - İLK 10 SATIR) ---")
print(y_ornek_df.to_string(index=False))

# =========================================================
# 12.2) EXCEL DOSYASI OLARAK KAYDET VE DÜZENLE
# =========================================================
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

# Önce dosyaları oluştur
X_hazir_df.to_excel("model_giris_X_tam.xlsx", index=False)
print(X_hazir_df.shape)
X_ornek_df.to_excel("model_giris_X_ornek.xlsx", index=False)
y_df.to_excel("model_giris_y_tam.xlsx", index=False)
y_ornek_df.to_excel("model_giris_y_ornek.xlsx", index=False)

# Excel dosyasını daha okunur hale getiren fonksiyon
def excel_duzenle(dosya_adi):
    wb = load_workbook(dosya_adi)
    ws = wb.active

    # Başlık satırını biçimlendir
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.fill = PatternFill(fill_type="solid", fgColor="D9EAF7")

    # Tüm hücreleri ortala
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center")

    # Sütun genişliklerini ayarla
    for col_cells in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col_cells[0].column)

        for cell in col_cells:
            try:
                if cell.value is not None:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass

        ws.column_dimensions[col_letter].width = min(max_length + 2, 25)

    # İlk satırı sabitle
    ws.freeze_panes = "A2"

    # Filtre ekle
    ws.auto_filter.ref = ws.dimensions

    wb.save(dosya_adi)

# Dosyaları düzenle
excel_duzenle("model_giris_X_tam.xlsx")
excel_duzenle("model_giris_X_ornek.xlsx")
excel_duzenle("model_giris_y_tam.xlsx")
excel_duzenle("model_giris_y_ornek.xlsx")

# =========================================================
# 13) MODELLERİ TANIMLA
# =========================================================
models = {
    "Linear Regression": LinearRegression(),
    "Decision Tree": DecisionTreeRegressor(random_state=42, max_depth=8),
    "Random Forest": RandomForestRegressor(
        n_estimators=200,
        max_depth=8,
        random_state=42
    ),
    "SVM": SVR(kernel="rbf")
}

results = []
trained_models = {}

# =========================================================
# 14) MODELLERİ EĞİT VE KARŞILAŞTIR
# =========================================================
for model_name, regressor in models.items():
    pipeline = Pipeline(steps=[
        ("preprocessor", preprocessor),
        ("regressor", regressor)
    ])

    # Model eğitimi
    pipeline.fit(X_train, y_train)

    # Tahminler
    y_pred_train = pipeline.predict(X_train)
    y_pred_test = pipeline.predict(X_test)

    # RMSE hesaplama
    train_rmse = np.sqrt(mean_squared_error(y_train, y_pred_train))
    test_rmse = np.sqrt(mean_squared_error(y_test, y_pred_test))

    # MAE hesaplama (EKLEDİĞİN EN KRİTİK KISIM)
    train_mae = mean_absolute_error(y_train, y_pred_train)
    test_mae = mean_absolute_error(y_test, y_pred_test)

    # R² hesaplama
    train_r2 = r2_score(y_train, y_pred_train)
    test_r2 = r2_score(y_test, y_pred_test)

    # Sonuçları tabloya ekleme
    results.append({
        "Model": model_name,
        "Egitim_RMSE": train_rmse,
        "Test_RMSE": test_rmse,
        "Egitim_MAE": train_mae,   # 🔥 EKLENEN
        "Test_MAE": test_mae,
        "Egitim_R2": train_r2,
        "Test_R2": test_r2
    })

    # Model ve tahminleri sakla
    trained_models[model_name] = {
        "pipeline": pipeline,
        "y_pred_test": y_pred_test
    }

# =========================================================
# 15) MODEL KARŞILAŞTIRMA TABLOSU
# =========================================================
results_df = pd.DataFrame(results)
results_df = results_df.sort_values(by="Test_R2", ascending=False).reset_index(drop=True)

print("\n--- MODEL KARŞILAŞTIRMA TABLOSU ---")
print(results_df.round(3).to_string(index=False))

# =========================================================
# 16) EN İYİ MODELİ SEÇ
# =========================================================
best_model_name = results_df.loc[0, "Model"]
best_model = trained_models[best_model_name]["pipeline"]
best_y_pred_test = trained_models[best_model_name]["y_pred_test"]

print("\n--- EN İYİ MODEL ---")
print(f"En iyi model: {best_model_name}")

print("\n--- EN İYİ MODEL YORUMU ---")
best_train_r2 = results_df.loc[0, "Egitim_R2"]
best_test_r2 = results_df.loc[0, "Test_R2"]

if abs(best_train_r2 - best_test_r2) < 0.10:
    print("Seçilen model eğitim ve test verisinde benzer performans göstermektedir. Genelleme başarısı kabul edilebilir.")
elif best_train_r2 > best_test_r2:
    print("Seçilen model eğitim verisinde daha iyi performans göstermektedir. Aşırı öğrenme olabilir.")
else:
    print("Seçilen model test verisinde beklenmedik şekilde farklı sonuç vermektedir. Veri yapısı kontrol edilmelidir.")

# =========================================================
# 17) HER MODEL İÇİN GERÇEK - TAHMİN GRAFİĞİ
# =========================================================
for model_name, model_data in trained_models.items():

    y_pred = model_data["y_pred_test"]

    plt.figure(figsize=(6, 6))
    plt.scatter(y_test, y_pred)

    min_v = min(y_test.min(), y_pred.min())
    max_v = max(y_test.max(), y_pred.max())

    plt.plot([min_v, max_v], [min_v, max_v], linestyle="--")

    plt.xlabel("Gerçek Başarı")
    plt.ylabel("Tahmin Edilen Başarı")
    plt.title(f"Gerçek - Tahmin Karşılaştırması ({model_name})")

    plt.tight_layout()
    plt.show()

# =========================================================
# 18) HİPERPARAMETRE OPTİMİZASYONU
# =========================================================
from sklearn.model_selection import GridSearchCV  # grid search ile en iyi parametreyi bulmak için

print("\n--- HİPERPARAMETRE OPTİMİZASYONU BAŞLADI ---")


# Verilen model için train/test metriklerini tek seferde hesaplayan yardımcı fonksiyon
def modeli_degerlendir(model, X_train, X_test, y_train, y_test):
    y_pred_train = model.predict(X_train)  # eğitim verisi tahmini
    y_pred_test = model.predict(X_test)    # test verisi tahmini

    train_rmse = np.sqrt(mean_squared_error(y_train, y_pred_train))  # eğitim RMSE
    test_rmse = np.sqrt(mean_squared_error(y_test, y_pred_test))     # test RMSE

    train_mae = mean_absolute_error(y_train, y_pred_train)  # eğitim MAE
    test_mae = mean_absolute_error(y_test, y_pred_test)     # test MAE

    train_r2 = r2_score(y_train, y_pred_train)  # eğitim R2
    test_r2 = r2_score(y_test, y_pred_test)     # test R2

    return {
        "Egitim_RMSE": train_rmse,
        "Test_RMSE": test_rmse,
        "Egitim_MAE": train_mae,
        "Test_MAE": test_mae,
        "Egitim_R2": train_r2,
        "Test_R2": test_r2,
        "y_pred_test": y_pred_test
    }


# =========================================================
# 18.1) LINEAR REGRESSION OPTİMİZASYONU
# =========================================================
lr_pipeline_opt = Pipeline(steps=[
    ("preprocessor", preprocessor),
    ("regressor", LinearRegression())
])

# Linear Regression için optimize edilecek parametre az olduğu için küçük grid yeterli
lr_param_grid = {
    "regressor__fit_intercept": [True, False]
}

lr_grid = GridSearchCV(
    estimator=lr_pipeline_opt,
    param_grid=lr_param_grid,
    cv=5,
    scoring="r2",
    n_jobs=-1
)

lr_grid.fit(X_train, y_train)

print("\n--- OPTİMİZE LINEAR REGRESSION EN İYİ PARAMETRELER ---")
print(lr_grid.best_params_)

print("\n--- OPTİMİZE LINEAR REGRESSION CV SKORU ---")
print(round(lr_grid.best_score_, 3))

best_lr_opt = lr_grid.best_estimator_
lr_opt_metrics = modeli_degerlendir(best_lr_opt, X_train, X_test, y_train, y_test)


# =========================================================
# 18.2) DECISION TREE OPTİMİZASYONU
# =========================================================
dt_pipeline_opt = Pipeline(steps=[
    ("preprocessor", preprocessor),
    ("regressor", DecisionTreeRegressor(random_state=42))
])

dt_param_grid = {
    "regressor__max_depth": [3, 5, 8, 10, None],
    "regressor__min_samples_split": [2, 5, 10, 15],
    "regressor__min_samples_leaf": [1, 2, 4, 6]
}

dt_grid = GridSearchCV(
    estimator=dt_pipeline_opt,
    param_grid=dt_param_grid,
    cv=5,
    scoring="r2",
    n_jobs=-1
)

dt_grid.fit(X_train, y_train)

print("\n--- OPTİMİZE DECISION TREE EN İYİ PARAMETRELER ---")
print(dt_grid.best_params_)

print("\n--- OPTİMİZE DECISION TREE CV SKORU ---")
print(round(dt_grid.best_score_, 3))

best_dt_opt = dt_grid.best_estimator_
dt_opt_metrics = modeli_degerlendir(best_dt_opt, X_train, X_test, y_train, y_test)


# =========================================================
# 18.3) RANDOM FOREST OPTİMİZASYONU
# =========================================================
rf_pipeline_opt = Pipeline(steps=[
    ("preprocessor", preprocessor),
    ("regressor", RandomForestRegressor(random_state=42))
])

rf_param_grid = {
    "regressor__n_estimators": [100, 200, 300],
    "regressor__max_depth": [5, 8, 10, None],
    "regressor__min_samples_split": [2, 5, 10],
    "regressor__min_samples_leaf": [1, 2, 4]
}

rf_grid = GridSearchCV(
    estimator=rf_pipeline_opt,
    param_grid=rf_param_grid,
    cv=5,
    scoring="r2",
    n_jobs=-1
)

rf_grid.fit(X_train, y_train)

print("\n--- OPTİMİZE RANDOM FOREST EN İYİ PARAMETRELER ---")
print(rf_grid.best_params_)

print("\n--- OPTİMİZE RANDOM FOREST CV SKORU ---")
print(round(rf_grid.best_score_, 3))

best_rf_opt = rf_grid.best_estimator_
rf_opt_metrics = modeli_degerlendir(best_rf_opt, X_train, X_test, y_train, y_test)

# =========================================================
# 18.4) SVM OPTİMİZASYONU
# =========================================================
svm_pipeline_opt = Pipeline(steps=[
    ("preprocessor", preprocessor),
    ("regressor", SVR())
])

svm_param_grid = {
    "regressor__kernel": ["rbf", "linear"],
    "regressor__C": [0.1, 1, 10, 100],
    "regressor__epsilon": [0.1, 0.5, 1],
    "regressor__gamma": ["scale", "auto"]
}

svm_grid = GridSearchCV(
    estimator=svm_pipeline_opt,
    param_grid=svm_param_grid,
    cv=5,
    scoring="r2",
    n_jobs=-1
)

svm_grid.fit(X_train, y_train)

print("\n--- OPTİMİZE SVM EN İYİ PARAMETRELER ---")
print(svm_grid.best_params_)

print("\n--- OPTİMİZE SVM CV SKORU ---")
print(round(svm_grid.best_score_, 3))

best_svm_opt = svm_grid.best_estimator_
svm_opt_metrics = modeli_degerlendir(best_svm_opt, X_train, X_test, y_train, y_test)

# =========================================================
# 19) OPTİMİZASYON SONRASI MODEL KARŞILAŞTIRMA
# =========================================================
opt_results_df = pd.DataFrame([
    {
        "Model": "Linear Regression",
        "Egitim_RMSE": lr_opt_metrics["Egitim_RMSE"],
        "Test_RMSE": lr_opt_metrics["Test_RMSE"],
        "Egitim_MAE": lr_opt_metrics["Egitim_MAE"],
        "Test_MAE": lr_opt_metrics["Test_MAE"],
        "Egitim_R2": lr_opt_metrics["Egitim_R2"],
        "Test_R2": lr_opt_metrics["Test_R2"]
    },
    {
        "Model": "Decision Tree",
        "Egitim_RMSE": dt_opt_metrics["Egitim_RMSE"],
        "Test_RMSE": dt_opt_metrics["Test_RMSE"],
        "Egitim_MAE": dt_opt_metrics["Egitim_MAE"],
        "Test_MAE": dt_opt_metrics["Test_MAE"],
        "Egitim_R2": dt_opt_metrics["Egitim_R2"],
        "Test_R2": dt_opt_metrics["Test_R2"]
    },
    {
        "Model": "Random Forest",
        "Egitim_RMSE": rf_opt_metrics["Egitim_RMSE"],
        "Test_RMSE": rf_opt_metrics["Test_RMSE"],
        "Egitim_MAE": rf_opt_metrics["Egitim_MAE"],
        "Test_MAE": rf_opt_metrics["Test_MAE"],
        "Egitim_R2": rf_opt_metrics["Egitim_R2"],
        "Test_R2": rf_opt_metrics["Test_R2"]
    },
    {
        "Model": "SVM",
        "Egitim_RMSE": svm_opt_metrics["Egitim_RMSE"],
        "Test_RMSE": svm_opt_metrics["Test_RMSE"],
        "Egitim_MAE": svm_opt_metrics["Egitim_MAE"],
        "Test_MAE": svm_opt_metrics["Test_MAE"],
        "Egitim_R2": svm_opt_metrics["Egitim_R2"],
        "Test_R2": svm_opt_metrics["Test_R2"]
    }
])

print("\n--- OPTİMİZASYON SONRASI MODEL KARŞILAŞTIRMA ---")
print(opt_results_df.round(3).to_string(index=False))


# =========================================================
# 20) ESKİ VE YENİ SONUÇLARI AYNI TABLODA TOPLA
# =========================================================
# results_df = optimizesiz tablo (önceden oluşturulmuştu)
# opt_results_df = optimize edilmiş tablo

old_table = results_df.copy()
old_table["Durum"] = "Optimizesiz"

new_table = opt_results_df.copy()
new_table["Durum"] = "Optimize"

karsilastirma_df = pd.concat([old_table, new_table], ignore_index=True)

# Sütun sıralaması
karsilastirma_df = karsilastirma_df[
    ["Model", "Durum", "Egitim_RMSE", "Test_RMSE", "Egitim_MAE", "Test_MAE", "Egitim_R2", "Test_R2"]
]

# Önce model adına göre, sonra durum bilgisine göre sırala
model_sirasi = ["Linear Regression", "Decision Tree", "Random Forest", "SVM"]
durum_sirasi = ["Optimizesiz", "Optimize"]

karsilastirma_df["Model"] = pd.Categorical(
    karsilastirma_df["Model"],
    categories=model_sirasi,
    ordered=True
)

karsilastirma_df["Durum"] = pd.Categorical(
    karsilastirma_df["Durum"],
    categories=durum_sirasi,
    ordered=True
)

karsilastirma_df = karsilastirma_df.sort_values(["Model", "Durum"]).reset_index(drop=True)

print("\n--- ESKİ VE YENİ MODELLERİN BİRLİKTE KARŞILAŞTIRILMASI ---")
print(karsilastirma_df.round(3).to_string(index=False))


# =========================================================
# 21) TABLOYU GRAFİK PENCERESİNDE DAHA DÜZGÜN GÖSTER
# =========================================================
tablo_gosterim = karsilastirma_df.round(3).copy()

# Kategorik tipleri tekrar stringe çevir ki tabloda düzgün gözüksün
tablo_gosterim["Model"] = tablo_gosterim["Model"].astype(str)
tablo_gosterim["Durum"] = tablo_gosterim["Durum"].astype(str)

fig, ax = plt.subplots(figsize=(14, 4.5))
ax.axis("off")

table = ax.table(
    cellText=tablo_gosterim.values,
    colLabels=tablo_gosterim.columns,
    loc="center",
    cellLoc="center"
)

table.auto_set_font_size(False)
table.set_fontsize(9)
table.scale(1.15, 1.5)

plt.title("Optimizesiz ve Optimize Modellerin Karşılaştırma Tablosu", pad=15)
plt.tight_layout()
plt.show()


# =========================================================
# 22) EN İYİ OPTİMİZE MODELİ SEÇ
# =========================================================
best_opt_row = opt_results_df.sort_values(by="Test_R2", ascending=False).reset_index(drop=True).loc[0]
best_opt_model_name = best_opt_row["Model"]

if best_opt_model_name == "Linear Regression":
    best_opt_model = best_lr_opt
    best_opt_test_pred = lr_opt_metrics["y_pred_test"]
    best_opt_train_r2 = lr_opt_metrics["Egitim_R2"]
    best_opt_test_r2 = lr_opt_metrics["Test_R2"]

elif best_opt_model_name == "Decision Tree":
    best_opt_model = best_dt_opt
    best_opt_test_pred = dt_opt_metrics["y_pred_test"]
    best_opt_train_r2 = dt_opt_metrics["Egitim_R2"]
    best_opt_test_r2 = dt_opt_metrics["Test_R2"]

elif best_opt_model_name == "SVM":
    best_opt_model = best_svm_opt
    best_opt_test_pred = svm_opt_metrics["y_pred_test"]
    best_opt_train_r2 = svm_opt_metrics["Egitim_R2"]
    best_opt_test_r2 = svm_opt_metrics["Test_R2"]

else:
    best_opt_model = best_rf_opt
    best_opt_test_pred = rf_opt_metrics["y_pred_test"]
    best_opt_train_r2 = rf_opt_metrics["Egitim_R2"]
    best_opt_test_r2 = rf_opt_metrics["Test_R2"]

print("\n--- EN İYİ OPTİMİZE MODEL ---")
print(best_opt_model_name)

print("\n--- EN İYİ OPTİMİZE MODEL YORUMU ---")
if abs(best_opt_train_r2 - best_opt_test_r2) < 0.10:
    print("Optimize edilen model eğitim ve test sonuçlarında daha dengeli görünmektedir.")
elif best_opt_train_r2 > best_opt_test_r2:
    print("Optimize sonrası da bir miktar overfitting eğilimi görülebilir.")
else:
    print("Optimize modelin test performansı beklenenden farklı görünmektedir.")


# =========================================================
# 23) EN İYİ OPTİMİZE MODEL İÇİN GERÇEK - TAHMİN GRAFİĞİ
# =========================================================
plt.figure(figsize=(6, 6))
plt.scatter(y_test, best_opt_test_pred)

min_v = min(y_test.min(), best_opt_test_pred.min())
max_v = max(y_test.max(), best_opt_test_pred.max())

plt.plot([min_v, max_v], [min_v, max_v], linestyle="--")

plt.xlabel("Gerçek Başarı")
plt.ylabel("Tahmin Edilen Başarı")
plt.title(f"Gerçek - Tahmin Karşılaştırması ({best_opt_model_name} - Optimize)")

plt.tight_layout()
plt.show()