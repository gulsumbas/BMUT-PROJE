import pandas as pd
import numpy as np
import matplotlib.pyplot as plt

from sklearn.model_selection import train_test_split
from sklearn.compose import ColumnTransformer
from sklearn.pipeline import Pipeline
from sklearn.preprocessing import OneHotEncoder, StandardScaler
from sklearn.impute import SimpleImputer

from sklearn.linear_model import LinearRegression
from sklearn.tree import DecisionTreeRegressor
from sklearn.ensemble import RandomForestRegressor
from sklearn.svm import SVR

from sklearn.metrics import mean_squared_error, mean_absolute_error, r2_score
from sklearn.feature_selection import RFE, mutual_info_regression

GRAFIK_AKTIF = False

df = pd.read_csv("anket_verisi.csv")

rename_map = {
    "Zaman damgası": "Zaman",
    "Cinsiyetiniz nedir?": "Cinsiyet",
    "Kaçıncı sınıfta eğitim görüyorsunuz?": "Sinif",
    "Hangi fakültede öğrenim görüyorsunuz?": "Fakulte",
    "Not ortalamanız hangi sistem üzerinden hesaplanıyor?": "Not_Sistemi",
    "4'lük sistemde geçen dönemki ortalamanızı giriniz (Küsüratlı değerleri virgül yerine noktayla ayırınız. Örn: 2.15):": "Gecen_4luk",
    "4'lük sistemde genel ortalamanızı giriniz (Küsüratlı değerleri virgül yerine noktayla ayırınız. Örn: 2.15):": "Genel_4luk",
    "100'lük sistemde geçen dönemki ortalamanızı giriniz (Küsüratlı değerleri virgül yerine noktayla ayırınız. Örn: 93.5):": "Gecen_100luk",
    "100'lük sistemde genel ortalamanızı giriniz (Küsüratlı değerleri virgül yerine noktayla ayırınız. Örn: 93.5):": "Genel_100luk",
    "Önceki dönem aldığınız ders sayısını giriniz:": "Ders_Sayisi",
    "Önceki dönem başarısız olduğunuz ders sayısını giriniz:": "Basarisiz_Ders",
    "Haftada kaç saat ders çalışıyorsunuz?": "Haftalik_Calisma",
    "Sınav haftaları kaç saat ders çalışıyorsunuz?": "Sinav_Haftasi_Calisma",
    "Annenizin eğitim durumu nedir?": "Anne_Egitim",
    "Babanızın eğitim durumu nedir?": "Baba_Egitim",
    "Evinizin aylık toplam geliri ne kadardır?": "Aylik_Gelir",
    "Bir işte çalışıyor musunuz?": "Calisma_Durumu",
    "Eğitim-öğretim dönemi içerisinde nerede ikamet ediyorsunuz?": "Ikamet",
    "Günlük ortalama ekran sürenizi seçiniz:": "Ekran_Suresi",
    "Günlük ortalama uyku sürenizi seçiniz:": "Uyku_Suresi"
}

df.rename(columns=rename_map, inplace=True)

def temizle_sayi(x):
    if pd.isna(x):
        return np.nan
    s = str(x).strip()
    if s == "":
        return np.nan
    if "," in s and "." in s:
        if s.rfind(",") > s.rfind("."):
            s = s.replace(".", "")
            s = s.replace(",", ".")
        else:
            s = s.replace(",", "")
    else:
        s = s.replace(",", ".")
    try:
        deger = float(s)

        # Eğer 4'lük sistem notu yanlışlıkla 228 gibi geldiyse
        # 2.28'e dönüştür
        if deger > 4 and deger <= 400:
            deger = deger / 100

        return deger
    except:
        return np.nan

sayisal_sutunlar = [
    "Gecen_4luk", "Genel_4luk", "Gecen_100luk", "Genel_100luk",
    "Ders_Sayisi", "Basarisiz_Ders"
]

for col in sayisal_sutunlar:

    df[col] = df[col].apply(temizle_sayi)

    # Sadece 4'lük sistem sütunlarında düzeltme uygula
    if col in ["Gecen_4luk", "Genel_4luk"]:

        df[col] = df[col].apply(
            lambda x: x / 100 if pd.notna(x) and x > 4 else x
        )

df["Ortak_Basari_100"] = np.where(
    df["Genel_100luk"].notna(), df["Genel_100luk"], df["Genel_4luk"] * 25
)
df["Ortak_Gecen_100"] = np.where(
    df["Gecen_100luk"].notna(), df["Gecen_100luk"], df["Gecen_4luk"] * 25
)
df["Ortak_Basari_4"] = np.where(
    df["Genel_4luk"].notna(), df["Genel_4luk"], df["Genel_100luk"] / 25
)

df.loc[(df["Ortak_Basari_100"] < 0) | (df["Ortak_Basari_100"] > 100), "Ortak_Basari_100"] = np.nan
df.loc[(df["Ortak_Gecen_100"] < 0) | (df["Ortak_Gecen_100"] > 100), "Ortak_Gecen_100"] = np.nan
df.loc[(df["Ortak_Basari_4"] < 0) | (df["Ortak_Basari_4"] > 4), "Ortak_Basari_4"] = np.nan

df = df[df["Ortak_Basari_100"].notna()].copy()

# =========================================================
# Mantıksız ders kayıtlarını temizle
# Başarısız ders sayısı toplam dersten fazla olamaz
# =========================================================
df.loc[
    df["Basarisiz_Ders"] > df["Ders_Sayisi"],
    "Basarisiz_Ders"
] = np.nan

def nadir_kategorileri_birlestir(df, col, min_count=10):
    counts = df[col].value_counts()
    rare = counts[counts < min_count].index
    df[col] = df[col].where(~df[col].isin(rare), other="Diğer")
    return df

for col in ["Fakulte", "Sinif", "Aylik_Gelir", "Ikamet"]:
    df = nadir_kategorileri_birlestir(df, col, min_count=10)

# =========================================================
# GERÇEK KATEGORİLERİ GÖSTER (map'leri düzeltmek için)
# =========================================================
print("\n--- GERÇEK KATEGORİLER ---")
for col in ["Sinif", "Aylik_Gelir", "Ekran_Suresi", "Uyku_Suresi",
            "Haftalik_Calisma", "Sinav_Haftasi_Calisma",
            "Anne_Egitim", "Baba_Egitim"]:
    print(f"\n{col}")
    print(df[col].value_counts(dropna=False))

print("\n--- VERİ SETİ GENEL BİLGİ ---")
print(df.info())

print("\n--- TOPLAM GEÇERLİ GÖZLEM SAYISI ---")
print(len(df))

print("\n--- NOT SİSTEMİ DAĞILIMI ---")
print(df["Not_Sistemi"].value_counts(dropna=False))

analiz_sayisal = [
    "Ortak_Basari_100", "Ortak_Gecen_100", "Ortak_Basari_4",
    "Ders_Sayisi", "Basarisiz_Ders"
]
analiz_sayisal = [c for c in analiz_sayisal if df[c].notna().sum() >= 2]

print("\n--- SAYISAL DEĞİŞKENLERİN TANIMSAL İSTATİSTİKLERİ ---")
print(df[analiz_sayisal].describe())

kategorik_sutunlar = [
    "Cinsiyet", "Sinif", "Fakulte", "Not_Sistemi",
    "Haftalik_Calisma", "Sinav_Haftasi_Calisma",
    "Anne_Egitim", "Baba_Egitim", "Aylik_Gelir",
    "Calisma_Durumu", "Ikamet", "Ekran_Suresi", "Uyku_Suresi"
]

print("\n--- KATEGORİK DEĞİŞKENLERİN FREKANSLARI ---")
for col in kategorik_sutunlar:
    print(f"\n{col}")
    print(df[col].value_counts(dropna=False))

print("\n--- KATEGORİK DEĞİŞKENLERE GÖRE ORTALAMA BAŞARI ---")
for col in kategorik_sutunlar:
    print(f"\n{col}")
    print(df.groupby(col)["Ortak_Basari_100"].agg(["count", "mean"]).sort_values("mean", ascending=False))

print("\n--- SAYISAL DEĞİŞKENLERİN BAŞARIYLA KORELASYONU ---")
corr_target = df[analiz_sayisal].corr(numeric_only=True)["Ortak_Basari_100"].sort_values(ascending=False)
print(corr_target)

if GRAFIK_AKTIF:
    df.groupby("Fakulte")["Ortak_Basari_100"].mean().sort_values().plot(kind="bar")
    plt.title("Fakülteye Göre Ortalama Başarı")
    plt.ylabel("Ortalama Başarı")
    plt.xticks(rotation=45)
    plt.tight_layout()
    plt.show()

if GRAFIK_AKTIF:
    plt.rcParams["figure.figsize"] = (10, 6)
    plt.rcParams["font.size"] = 10
    for column in df.columns:
        plt.figure()
        if pd.api.types.is_numeric_dtype(df[column]):
            temiz_veri = df[column].dropna()
            if len(temiz_veri) > 0:
                plt.hist(temiz_veri, bins=12, edgecolor="black")
                ortalama = temiz_veri.mean()
                medyan = temiz_veri.median()
                plt.axvline(ortalama, linestyle="--", linewidth=2, label=f"Ortalama: {ortalama:.2f}")
                plt.axvline(medyan, linestyle="-.", linewidth=2, label=f"Medyan: {medyan:.2f}")
                plt.title(f"{column} - Sayısal Dağılım", fontsize=13, fontweight="bold")
                plt.xlabel(column)
                plt.ylabel("Frekans")
                plt.grid(axis="y", linestyle="--", alpha=0.5)
                plt.legend()
            else:
                plt.text(0.5, 0.5, "Gösterilecek veri yok", ha="center", va="center")
                plt.title(f"{column} - Sayısal Dağılım")
        else:
            sayim = df[column].fillna("Boş").astype(str).value_counts()
            kategori_sayisi = len(sayim)
            if kategori_sayisi > 0:
                if kategori_sayisi <= 3:
                    plt.pie(sayim.values, labels=sayim.index, autopct="%1.1f%%", startangle=90)
                    plt.title(f"{column} - Oransal Dağılım", fontsize=13, fontweight="bold")
                    plt.axis("equal")
                else:
                    if kategori_sayisi > 12:
                        sayim = sayim.head(12)
                    ax = sayim.plot(kind="bar", edgecolor="black")
                    plt.title(f"{column} - Kategorik Dağılım", fontsize=13, fontweight="bold")
                    plt.xlabel(column)
                    plt.ylabel("Frekans")
                    plt.xticks(rotation=35, ha="right")
                    plt.grid(axis="y", linestyle="--", alpha=0.5)
                    for p in ax.patches:
                        yukseklik = p.get_height()
                        ax.annotate(str(int(yukseklik)),
                            (p.get_x() + p.get_width() / 2, yukseklik),
                            ha="center", va="bottom", fontsize=9)
            else:
                plt.text(0.5, 0.5, "Gösterilecek veri yok", ha="center", va="center")
                plt.title(f"{column} - Grafik")
        plt.tight_layout()
        plt.show()

corr_matrix = df[analiz_sayisal].corr(numeric_only=True)
print("\n--- KORELASYON MATRİSİ ---")
print(corr_matrix)

if GRAFIK_AKTIF:
    plt.figure(figsize=(8, 6))
    plt.imshow(corr_matrix, cmap="coolwarm", interpolation="nearest")
    plt.colorbar()
    plt.xticks(range(len(corr_matrix.columns)), corr_matrix.columns, rotation=45, ha="right")
    plt.yticks(range(len(corr_matrix.columns)), corr_matrix.columns)
    plt.title("Korelasyon Matrisi")
    plt.tight_layout()
    plt.show()

    plt.figure(figsize=(8, 5))
    df["Ortak_Basari_100"].hist(bins=10)
    plt.title("Genel Başarı Dağılımı")
    plt.xlabel("Başarı Notu (100'lük)")
    plt.ylabel("Frekans")
    plt.tight_layout()
    plt.show()

# =========================================================
# 9.5) ORDINAL DÖNÜŞÜMLER
# =========================================================

egitim_map = {
    "Hiçbiri": 0,
    "İlkokul": 1,
    "Ortaokul": 2,
    "Lise": 3,
    "Üniversite": 4
}

df["Anne_Egitim"] = df["Anne_Egitim"].map(egitim_map)
df["Baba_Egitim"] = df["Baba_Egitim"].map(egitim_map)

calisma_map = {
    "0-5 saat": 1,
    "5-10 saat": 2,
    "10-20 saat": 3,
    "20-30 saat": 4,
    "30-50 saat": 5,
    "50+ saat": 6
}

df["Haftalik_Calisma"] = df["Haftalik_Calisma"].map(calisma_map)
df["Sinav_Haftasi_Calisma"] = df["Sinav_Haftasi_Calisma"].map(calisma_map)

sinif_map = {
    "Hazırlık": 0,
    "1.sınıf": 1,
    "2.sınıf": 2,
    "3.sınıf": 3,
    "4.sınıf": 4,
    "Diğer": 3
}

df["Sinif"] = df["Sinif"].map(sinif_map)

gelir_map = {
    "0 - 28.075 TL": 1,
    "28.075 - 60.000 TL": 2,
    "60.000 - 90.000 TL": 3,
    "90.000+ TL": 4
}

df["Aylik_Gelir"] = df["Aylik_Gelir"].map(gelir_map)

ekran_map = {
    "0 - 3 saat": 1,
    "3 - 6 saat": 2,
    "7 - 10 saat": 3,
    "10 - 15 saat": 4,
    "15 - 24 saat": 5
}

df["Ekran_Suresi"] = df["Ekran_Suresi"].map(ekran_map)

uyku_map = {
    "0 - 3 saat": 1,
    "3 - 6 saat": 2,
    "7 - 10 saat": 3,
    "10 - 15 saat": 4,
    "15 - 24 saat": 5
}

df["Uyku_Suresi"] = df["Uyku_Suresi"].map(uyku_map)

print("\n--- ORDINAL DÖNÜŞÜM SONRASI BOŞ DEĞER KONTROLÜ ---")
print(df[[
    "Anne_Egitim",
    "Baba_Egitim",
    "Haftalik_Calisma",
    "Sinav_Haftasi_Calisma",
    "Sinif",
    "Aylik_Gelir",
    "Ekran_Suresi",
    "Uyku_Suresi"
]].isna().sum())

# =========================================================
# ÖNEMLİ: Aylik_Gelir, Ekran_Suresi, Uyku_Suresi
# map'lerini GERÇEK KATEGORİLER çıktısına göre doldur.
# Şimdilik kategorik olarak bırakıyoruz — hata vermesin diye.
# =========================================================

# =========================================================
# 9.6) FEATURE ENGINEERING
# =========================================================

# Başarısız ders oranı
df["Basarisiz_Orani"] = (
    df["Basarisiz_Ders"] /
    (df["Ders_Sayisi"] + 1)
)

# Sınav haftası çalışma artışı
df["Calisma_Artisi"] = (
    df["Sinav_Haftasi_Calisma"] -
    df["Haftalik_Calisma"]
)

# Ekran süresi / uyku dengesi
df["Ekran_Uyku_Orani"] = (
    df["Ekran_Suresi"] /
    (df["Uyku_Suresi"] + 1)
)

print("\n--- FEATURE ENGINEERING SONRASI YENİ DEĞİŞKENLER ---")
print(df[[
    "Basarisiz_Orani",
    "Calisma_Artisi",
    "Ekran_Uyku_Orani"
]].head())

# =========================================================
# 10) MODEL İÇİN VERİ HAZIRLAMA
# =========================================================
sayisal_ozellikler = [
    "Ortak_Gecen_100",
    "Ders_Sayisi",
    "Basarisiz_Ders",
    "Anne_Egitim",
    "Baba_Egitim",
    "Haftalik_Calisma",
    "Sinav_Haftasi_Calisma",
    "Sinif",
    "Aylik_Gelir",
    "Ekran_Suresi",
    "Uyku_Suresi",

    # Feature engineering ile üretilen yeni değişkenler
    "Basarisiz_Orani",
    "Calisma_Artisi",
    "Ekran_Uyku_Orani"
]

kategorik_ozellikler = [
    "Cinsiyet",
    "Fakulte",
    "Calisma_Durumu",
    "Ikamet"
]

X = df[sayisal_ozellikler + kategorik_ozellikler].copy()
y = df["Ortak_Basari_100"].copy()

X_train, X_test, y_train, y_test = train_test_split(
    X, y, test_size=0.20, random_state=42
)

print("\n--- EĞİTİM / TEST BOYUTLARI ---")
print("X_train:", X_train.shape)
print("X_test :", X_test.shape)

num_transformer = Pipeline(steps=[
    ("imputer", SimpleImputer(strategy="median")),
    ("scaler", StandardScaler())
])

cat_transformer = Pipeline(steps=[
    ("imputer", SimpleImputer(strategy="most_frequent")),
    ("onehot", OneHotEncoder(handle_unknown="ignore"))
])

preprocessor = ColumnTransformer(
    transformers=[
        ("num", num_transformer, sayisal_ozellikler),
        ("cat", cat_transformer, kategorik_ozellikler)
    ]
)

X_hazir = preprocessor.fit_transform(X)
if hasattr(X_hazir, "toarray"):
    X_hazir = X_hazir.toarray()

encoded_cat_features = preprocessor.named_transformers_["cat"]["onehot"].get_feature_names_out(kategorik_ozellikler)
tum_feature_isimleri = sayisal_ozellikler + list(encoded_cat_features)

X_hazir_df = pd.DataFrame(X_hazir, columns=tum_feature_isimleri)
y_df = pd.DataFrame(y.values, columns=["Ortak_Basari_100"])

X_ornek_df = X_hazir_df.head(10).copy()
y_ornek_df = y_df.head(10).copy()

print("\n--- MODELİN GÖRDÜĞÜ X MATRİSİ (İLK 10 SATIR) ---")
print(X_ornek_df.to_string(index=False))
print("\nToplam gözlem sayısı:", X_hazir_df.shape[0])
print("Toplam özellik (feature) sayısı:", X_hazir_df.shape[1])
print("\n--- MODELİN GÖRDÜĞÜ Y (HEDEF DEĞİŞKEN - İLK 10 SATIR) ---")
print(y_ornek_df.to_string(index=False))

rfe_rapor_model = LinearRegression()
rfe_rapor = RFE(estimator=rfe_rapor_model, n_features_to_select=10)
rfe_rapor.fit(X_hazir_df, y)

rfe_sonuclari_df = pd.DataFrame({
    "Feature": tum_feature_isimleri,
    "Secildi_Mi": rfe_rapor.support_,
    "Siralama": rfe_rapor.ranking_
}).sort_values(by=["Secildi_Mi", "Siralama"], ascending=[False, True])

print("\n--- RFE İLE SEÇİLEN FEATURE'LAR ---")
print(rfe_sonuclari_df[rfe_sonuclari_df["Secildi_Mi"] == True].to_string(index=False))

mi_skorlari = mutual_info_regression(X_hazir_df, y, random_state=42)
mi_sonuclari_df = pd.DataFrame({
    "Feature": tum_feature_isimleri,
    "MI_Skoru": mi_skorlari
}).sort_values(by="MI_Skoru", ascending=False).reset_index(drop=True)

print("\n--- MUTUAL INFORMATION FEATURE SKORLARI ---")
print(mi_sonuclari_df.head(15).to_string(index=False))

from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

X_hazir_df.to_excel("model_giris_X_tam.xlsx", index=False)
X_ornek_df.to_excel("model_giris_X_ornek.xlsx", index=False)
y_df.to_excel("model_giris_y_tam.xlsx", index=False)
y_ornek_df.to_excel("model_giris_y_ornek.xlsx", index=False)
rfe_sonuclari_df.to_excel("rfe_feature_secimi.xlsx", index=False)
mi_sonuclari_df.to_excel("mutual_information_skorlari.xlsx", index=False)

def excel_duzenle(dosya_adi):
    wb = load_workbook(dosya_adi)
    ws = wb.active
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center")
    for col_cells in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            try:
                if cell.value is not None:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[col_letter].width = min(max_length + 2, 35)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    wb.save(dosya_adi)

for f in ["model_giris_X_tam.xlsx", "model_giris_X_ornek.xlsx",
          "model_giris_y_tam.xlsx", "model_giris_y_ornek.xlsx",
          "rfe_feature_secimi.xlsx", "mutual_information_skorlari.xlsx"]:
    excel_duzenle(f)

models = {
    "Linear Regression": LinearRegression(),
    "Decision Tree": DecisionTreeRegressor(random_state=42, max_depth=8),
    "Random Forest": RandomForestRegressor(n_estimators=200, max_depth=8, random_state=42),
    "SVM": SVR(kernel="rbf")
}

results = []
trained_models = {}

for model_name, regressor in models.items():
    pipeline = Pipeline(steps=[
        ("preprocessor", preprocessor),
        ("feature_selection", RFE(estimator=LinearRegression(), n_features_to_select=15)),
        ("regressor", regressor)
    ])
    pipeline.fit(X_train, y_train)
    y_pred_train = pipeline.predict(X_train)
    y_pred_test = pipeline.predict(X_test)

    results.append({
        "Model": model_name,
        "Egitim_RMSE": np.sqrt(mean_squared_error(y_train, y_pred_train)),
        "Test_RMSE": np.sqrt(mean_squared_error(y_test, y_pred_test)),
        "Egitim_MAE": mean_absolute_error(y_train, y_pred_train),
        "Test_MAE": mean_absolute_error(y_test, y_pred_test),
        "Egitim_R2": r2_score(y_train, y_pred_train),
        "Test_R2": r2_score(y_test, y_pred_test)
    })
    trained_models[model_name] = {"pipeline": pipeline, "y_pred_test": y_pred_test}

results_df = pd.DataFrame(results).sort_values(by="Test_R2", ascending=False).reset_index(drop=True)
print("\n--- MODEL KARŞILAŞTIRMA TABLOSU ---")
print(results_df.round(3).to_string(index=False))

best_model_name = results_df.loc[0, "Model"]
best_model = trained_models[best_model_name]["pipeline"]
best_y_pred_test = trained_models[best_model_name]["y_pred_test"]

print("\n--- EN İYİ MODEL ---")
print(f"En iyi model: {best_model_name}")

best_train_r2 = results_df.loc[0, "Egitim_R2"]
best_test_r2 = results_df.loc[0, "Test_R2"]

print("\n--- EN İYİ MODEL YORUMU ---")
if abs(best_train_r2 - best_test_r2) < 0.10:
    print("Seçilen model eğitim ve test verisinde benzer performans göstermektedir.")
elif best_train_r2 > best_test_r2:
    print("Seçilen model eğitim verisinde daha iyi performans göstermektedir. Aşırı öğrenme olabilir.")
else:
    print("Seçilen model test verisinde beklenmedik şekilde farklı sonuç vermektedir.")

for model_name, model_data in trained_models.items():
    y_pred = model_data["y_pred_test"]
    plt.figure(figsize=(6, 6))
    plt.scatter(y_test, y_pred)
    min_v = min(y_test.min(), y_pred.min())
    max_v = max(y_test.max(), y_pred.max())
    plt.plot([min_v, max_v], [min_v, max_v], linestyle="--")
    plt.xlabel("Gerçek Başarı")
    plt.ylabel("Tahmin Edilen Başarı")
    plt.title(f"Gerçek - Tahmin Karşılaştırması ({model_name})")
    plt.tight_layout()
    plt.show()

from sklearn.model_selection import GridSearchCV

print("\n--- HİPERPARAMETRE OPTİMİZASYONU BAŞLADI ---")

def modeli_degerlendir(model, X_train, X_test, y_train, y_test):
    y_pred_train = model.predict(X_train)
    y_pred_test = model.predict(X_test)
    return {
        "Egitim_RMSE": np.sqrt(mean_squared_error(y_train, y_pred_train)),
        "Test_RMSE": np.sqrt(mean_squared_error(y_test, y_pred_test)),
        "Egitim_MAE": mean_absolute_error(y_train, y_pred_train),
        "Test_MAE": mean_absolute_error(y_test, y_pred_test),
        "Egitim_R2": r2_score(y_train, y_pred_train),
        "Test_R2": r2_score(y_test, y_pred_test),
        "y_pred_test": y_pred_test
    }

# Linear Regression
lr_pipeline_opt = Pipeline(steps=[
    ("preprocessor", preprocessor),
    ("feature_selection", RFE(estimator=LinearRegression(), n_features_to_select=15)),
    ("regressor", LinearRegression())
])
lr_grid = GridSearchCV(lr_pipeline_opt, {"regressor__fit_intercept": [True, False]},
                       cv=5, scoring="r2", n_jobs=-1)
lr_grid.fit(X_train, y_train)
print("\n--- OPTİMİZE LINEAR REGRESSION EN İYİ PARAMETRELER ---")
print(lr_grid.best_params_)
print("CV Skoru:", round(lr_grid.best_score_, 3))
best_lr_opt = lr_grid.best_estimator_
lr_opt_metrics = modeli_degerlendir(best_lr_opt, X_train, X_test, y_train, y_test)

# Decision Tree
dt_pipeline_opt = Pipeline(steps=[
    ("preprocessor", preprocessor),
    ("feature_selection", RFE(estimator=LinearRegression(), n_features_to_select=15)),
    ("regressor", DecisionTreeRegressor(random_state=42))
])
dt_param_grid = {
    "regressor__max_depth": [3, 5, 8, 10, None],
    "regressor__min_samples_split": [2, 5, 10, 15],
    "regressor__min_samples_leaf": [1, 2, 4, 6]
}
dt_grid = GridSearchCV(dt_pipeline_opt, dt_param_grid, cv=5, scoring="r2", n_jobs=-1)
dt_grid.fit(X_train, y_train)
print("\n--- OPTİMİZE DECISION TREE EN İYİ PARAMETRELER ---")
print(dt_grid.best_params_)
print("CV Skoru:", round(dt_grid.best_score_, 3))
best_dt_opt = dt_grid.best_estimator_
dt_opt_metrics = modeli_degerlendir(best_dt_opt, X_train, X_test, y_train, y_test)

# Random Forest
rf_pipeline_opt = Pipeline(steps=[
    ("preprocessor", preprocessor),
    ("feature_selection", RFE(estimator=LinearRegression(), n_features_to_select=15)),
    ("regressor", RandomForestRegressor(random_state=42))
])
rf_param_grid = {
    "regressor__n_estimators": [100, 200, 300],
    "regressor__max_depth": [5, 8, 10, None],
    "regressor__min_samples_split": [2, 5, 10],
    "regressor__min_samples_leaf": [1, 2, 4]
}
rf_grid = GridSearchCV(rf_pipeline_opt, rf_param_grid, cv=5, scoring="r2", n_jobs=-1)
rf_grid.fit(X_train, y_train)
print("\n--- OPTİMİZE RANDOM FOREST EN İYİ PARAMETRELER ---")
print(rf_grid.best_params_)
print("CV Skoru:", round(rf_grid.best_score_, 3))
best_rf_opt = rf_grid.best_estimator_
rf_opt_metrics = modeli_degerlendir(best_rf_opt, X_train, X_test, y_train, y_test)

# SVM
svm_pipeline_opt = Pipeline(steps=[
    ("preprocessor", preprocessor),
    ("feature_selection", RFE(estimator=LinearRegression(), n_features_to_select=15)),
    ("regressor", SVR())
])
svm_param_grid = {
    "regressor__kernel": ["rbf", "linear"],
    "regressor__C": [0.1, 1, 10, 100],
    "regressor__epsilon": [0.1, 0.5, 1],
    "regressor__gamma": ["scale", "auto"]
}
svm_grid = GridSearchCV(svm_pipeline_opt, svm_param_grid, cv=5, scoring="r2", n_jobs=-1)
svm_grid.fit(X_train, y_train)
print("\n--- OPTİMİZE SVM EN İYİ PARAMETRELER ---")
print(svm_grid.best_params_)
print("CV Skoru:", round(svm_grid.best_score_, 3))
best_svm_opt = svm_grid.best_estimator_
svm_opt_metrics = modeli_degerlendir(best_svm_opt, X_train, X_test, y_train, y_test)

opt_results_df = pd.DataFrame([
    {"Model": "Linear Regression", **{k: v for k, v in lr_opt_metrics.items() if k != "y_pred_test"}},
    {"Model": "Decision Tree",     **{k: v for k, v in dt_opt_metrics.items() if k != "y_pred_test"}},
    {"Model": "Random Forest",     **{k: v for k, v in rf_opt_metrics.items() if k != "y_pred_test"}},
    {"Model": "SVM",               **{k: v for k, v in svm_opt_metrics.items() if k != "y_pred_test"}}
])

print("\n--- OPTİMİZASYON SONRASI MODEL KARŞILAŞTIRMA ---")
print(opt_results_df.round(3).to_string(index=False))

old_table = results_df.copy()
old_table["Durum"] = "Optimizesiz"
new_table = opt_results_df.copy()
new_table["Durum"] = "Optimize"

karsilastirma_df = pd.concat([old_table, new_table], ignore_index=True)
karsilastirma_df = karsilastirma_df[
    ["Model", "Durum", "Egitim_RMSE", "Test_RMSE", "Egitim_MAE", "Test_MAE", "Egitim_R2", "Test_R2"]
]

model_sirasi = ["Linear Regression", "Decision Tree", "Random Forest", "SVM"]
durum_sirasi = ["Optimizesiz", "Optimize"]
karsilastirma_df["Model"] = pd.Categorical(karsilastirma_df["Model"], categories=model_sirasi, ordered=True)
karsilastirma_df["Durum"] = pd.Categorical(karsilastirma_df["Durum"], categories=durum_sirasi, ordered=True)
karsilastirma_df = karsilastirma_df.sort_values(["Model", "Durum"]).reset_index(drop=True)

print("\n--- ESKİ VE YENİ MODELLERİN BİRLİKTE KARŞILAŞTIRILMASI ---")
print(karsilastirma_df.round(3).to_string(index=False))

tablo_gosterim = karsilastirma_df.round(3).copy()
tablo_gosterim["Model"] = tablo_gosterim["Model"].astype(str)
tablo_gosterim["Durum"] = tablo_gosterim["Durum"].astype(str)

fig, ax = plt.subplots(figsize=(14, 4.5))
ax.axis("off")
table = ax.table(
    cellText=tablo_gosterim.values,
    colLabels=tablo_gosterim.columns,
    loc="center", cellLoc="center"
)
table.auto_set_font_size(False)
table.set_fontsize(9)
table.scale(1.15, 1.5)
plt.title("Optimizesiz ve Optimize Modellerin Karşılaştırma Tablosu", pad=15)
plt.tight_layout()
plt.show()

best_opt_row = opt_results_df.sort_values(by="Test_R2", ascending=False).reset_index(drop=True).loc[0]
best_opt_model_name = best_opt_row["Model"]

metrics_map = {
    "Linear Regression": (best_lr_opt, lr_opt_metrics),
    "Decision Tree": (best_dt_opt, dt_opt_metrics),
    "Random Forest": (best_rf_opt, rf_opt_metrics),
    "SVM": (best_svm_opt, svm_opt_metrics)
}
best_opt_model, best_opt_metrics = metrics_map[best_opt_model_name]
best_opt_test_pred = best_opt_metrics["y_pred_test"]
best_opt_train_r2 = best_opt_metrics["Egitim_R2"]
best_opt_test_r2 = best_opt_metrics["Test_R2"]

print("\n--- EN İYİ OPTİMİZE MODEL ---")
print(best_opt_model_name)

print("\n--- EN İYİ OPTİMİZE MODEL YORUMU ---")
if abs(best_opt_train_r2 - best_opt_test_r2) < 0.10:
    print("Optimize edilen model eğitim ve test sonuçlarında daha dengeli görünmektedir.")
elif best_opt_train_r2 > best_opt_test_r2:
    print("Optimize sonrası da bir miktar overfitting eğilimi görülebilir.")
else:
    print("Optimize modelin test performansı beklenenden farklı görünmektedir.")

plt.figure(figsize=(6, 6))
plt.scatter(y_test, best_opt_test_pred)
min_v = min(y_test.min(), best_opt_test_pred.min())
max_v = max(y_test.max(), best_opt_test_pred.max())
plt.plot([min_v, max_v], [min_v, max_v], linestyle="--")
plt.xlabel("Gerçek Başarı")
plt.ylabel("Tahmin Edilen Başarı")
plt.title(f"Gerçek - Tahmin Karşılaştırması ({best_opt_model_name} - Optimize)")
plt.tight_layout()
plt.show()